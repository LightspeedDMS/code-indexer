"""QueryAnalyticsExportService -- export search event log to Excel (Issue #1160).

Provides dual-backend storage (SQLite/Postgres) for export history tracking,
and an export service that queries the search_event_log, writes an Excel file,
and records the result in the query_analytics_exports table.

Follows the SearchEventLogWriter pattern for dual-backend design.
"""

import logging
import os
import sqlite3
import time
import uuid as _uuid_module
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, cast

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Excel column names in exact canonical order (spec)
_EXCEL_COLUMNS = [
    "timestamp",
    "user",
    "repo_alias",
    "search_type",
    "query_text",
    "voyage_cache_hit",
    "voyage_cache_mode",
    "voyage_latency_ms",
    "cohere_cache_hit",
    "cohere_cache_mode",
    "cohere_latency_ms",
    "total_latency_ms",
    "result_count",
    "node_id",
    "correlation_id",
]

# Max rows per export (anti-unbounded-loop -- Messi #14)
_EXPORT_LIMIT = 50_000

# Number of seconds in one day
_SECONDS_PER_DAY = 86_400

# SQLite connection timeout
_SQLITE_TIMEOUT_SECONDS = 30

# Allowed fields for update_export (defence against SQL injection via dynamic SET)
_UPDATE_ALLOWED = frozenset(
    {
        "status",
        "file_path",
        "file_size_bytes",
        "row_count",
        "error_message",
        "retention_until",
    }
)

# Staleness threshold for orphan reconciliation (Bug #1228).
# Exports still in pending/running after this many seconds are considered orphaned.
# A real export (50k-row DB query + Excel write) completes in seconds to ~2 minutes.
# 300 seconds (5 min) gives ample margin for legitimate in-flight exports on other
# cluster nodes while definitively catching hours-old stuck exports from node deaths.
_DEFAULT_ORPHAN_THRESHOLD_SECONDS = 300.0

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _validate_export_id(export_id: str) -> None:
    """Raise ValueError if export_id is not a valid UUID string.

    Using uuid.UUID() is the canonical way to validate UUID format and prevents
    path traversal via export_id (e.g. '../../etc/passwd').
    """
    try:
        parsed = _uuid_module.UUID(export_id)
        if str(parsed) != export_id.lower():
            raise ValueError("non-canonical UUID form")
    except (ValueError, AttributeError) as exc:
        raise ValueError(
            f"export_id must be a canonical UUID string, got {export_id!r}"
        ) from exc


def _build_record_tuple(record: Dict[str, Any]) -> tuple:
    """Map a record dict to the canonical INSERT column order."""
    return (
        record["id"],
        record["initiated_by"],
        record["created_at"],
        record["status"],
        record["filter_summary"],
        record.get("file_path"),
        record.get("file_size_bytes"),
        record.get("row_count"),
        record.get("error_message"),
        record.get("retention_until"),
    )


def _validate_update_fields(fields: Dict[str, Any]) -> None:
    for key in fields:
        if key not in _UPDATE_ALLOWED:
            raise ValueError(f"update_export: unknown field {key!r}")


def _delete_expired_rows(rows: list, conn: Any, placeholder: str) -> int:
    """Remove files and DELETE DB rows for each expired export. Returns count."""
    count = 0
    for row in rows:
        file_path = row["file_path"] if hasattr(row, "keys") else row[1]
        if file_path:
            try:
                os.remove(file_path)
            except FileNotFoundError:
                pass
            except Exception as exc:
                logger.warning(
                    "evict_old_exports: could not remove %s: %s", file_path, exc
                )
        row_id = row["id"] if hasattr(row, "keys") else row[0]
        conn.execute(
            f"DELETE FROM query_analytics_exports WHERE id = {placeholder}",
            (row_id,),
        )
        count += 1
    return count


# ---------------------------------------------------------------------------
# Table DDL (shared, parameterised on timestamp type)
# ---------------------------------------------------------------------------

_CREATE_TABLE_SQLITE = """
    CREATE TABLE IF NOT EXISTS query_analytics_exports (
        id              TEXT PRIMARY KEY,
        initiated_by    TEXT NOT NULL,
        created_at      REAL NOT NULL,
        status          TEXT NOT NULL,
        filter_summary  TEXT NOT NULL,
        file_path       TEXT,
        file_size_bytes INTEGER,
        row_count       INTEGER,
        error_message   TEXT,
        retention_until REAL
    )
"""

_CREATE_TABLE_POSTGRES = """
    CREATE TABLE IF NOT EXISTS query_analytics_exports (
        id              TEXT PRIMARY KEY,
        initiated_by    TEXT NOT NULL,
        created_at      DOUBLE PRECISION NOT NULL,
        status          TEXT NOT NULL,
        filter_summary  TEXT NOT NULL,
        file_path       TEXT,
        file_size_bytes INTEGER,
        row_count       INTEGER,
        error_message   TEXT,
        retention_until DOUBLE PRECISION
    )
"""

_INSERT_SQL_SQLITE = """
    INSERT INTO query_analytics_exports
        (id, initiated_by, created_at, status, filter_summary,
         file_path, file_size_bytes, row_count, error_message, retention_until)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
"""

_INSERT_SQL_POSTGRES = """
    INSERT INTO query_analytics_exports
        (id, initiated_by, created_at, status, filter_summary,
         file_path, file_size_bytes, row_count, error_message, retention_until)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
"""

# ---------------------------------------------------------------------------
# SQLite backend
# ---------------------------------------------------------------------------


class QueryAnalyticsExportSqliteBackend:
    """SQLite backend for query_analytics_exports table (solo/dev mode)."""

    def __init__(self, db_path: str) -> None:
        self._db_path = db_path
        self._ensure_schema()

    def _ensure_schema(self) -> None:
        try:
            conn = sqlite3.connect(self._db_path, timeout=_SQLITE_TIMEOUT_SECONDS)
            try:
                conn.execute("PRAGMA journal_mode = WAL")
                conn.execute(_CREATE_TABLE_SQLITE)
                conn.commit()
            finally:
                conn.close()
        except Exception as exc:
            logger.warning(
                "QueryAnalyticsExportSqliteBackend: schema setup failed: %s", exc
            )

    def _conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self._db_path, timeout=_SQLITE_TIMEOUT_SECONDS)
        conn.row_factory = sqlite3.Row
        return conn

    def create_export(self, record: Dict[str, Any]) -> None:
        conn = self._conn()
        try:
            conn.execute(_INSERT_SQL_SQLITE, _build_record_tuple(record))
            conn.commit()
        finally:
            conn.close()

    def update_export(self, export_id: str, **fields: Any) -> None:
        if not fields:
            return
        _validate_update_fields(fields)
        set_clauses = [f"{k} = ?" for k in fields]
        params: List[Any] = list(fields.values()) + [export_id]
        conn = self._conn()
        try:
            conn.execute(
                f"UPDATE query_analytics_exports SET {', '.join(set_clauses)} WHERE id = ?",
                params,
            )
            conn.commit()
        finally:
            conn.close()

    def list_exports(self, export_id: Optional[str] = None) -> List[Dict[str, Any]]:
        conn = self._conn()
        try:
            if export_id is not None:
                rows = conn.execute(
                    "SELECT * FROM query_analytics_exports WHERE id = ? "
                    "ORDER BY created_at DESC",
                    (export_id,),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM query_analytics_exports ORDER BY created_at DESC"
                ).fetchall()
            return [dict(row) for row in rows]
        finally:
            conn.close()

    def evict_old_exports(self, now_ts: float) -> int:
        """Delete expired exports (retention_until < now_ts) and their files."""
        conn = self._conn()
        try:
            rows = conn.execute(
                "SELECT id, file_path FROM query_analytics_exports "
                "WHERE retention_until IS NOT NULL AND retention_until < ?",
                (now_ts,),
            ).fetchall()
            count = _delete_expired_rows(rows, conn, "?")
            conn.commit()
            return count
        finally:
            conn.close()

    def reconcile_orphaned_exports(
        self,
        threshold_seconds: float = _DEFAULT_ORPHAN_THRESHOLD_SECONDS,
        error: str = "interrupted by worker restart",
    ) -> int:
        """Mark stale pending/running exports as failed (Bug #1228).

        Cluster-safe predicate: only affects exports whose created_at is older
        than threshold_seconds ago.  A legitimately-running export on another
        cluster node (started seconds/minutes ago) is NOT within the orphan
        window.  Exports that have been stuck for hours (e.g. after an NFS
        outage that killed the owning worker) are definitively caught.

        Returns the number of rows transitioned to 'failed'.
        """
        cutoff_ts = time.time() - threshold_seconds
        conn = self._conn()
        try:
            cur = conn.execute(
                "UPDATE query_analytics_exports "
                "SET status = 'failed', error_message = ? "
                "WHERE status IN ('pending', 'running') AND created_at < ?",
                (error, cutoff_ts),
            )
            conn.commit()
            return cur.rowcount
        finally:
            conn.close()


# ---------------------------------------------------------------------------
# PostgreSQL backend
# ---------------------------------------------------------------------------


class QueryAnalyticsExportPostgresBackend:
    """PostgreSQL backend for query_analytics_exports table (cluster mode)."""

    def __init__(self, pool: Any) -> None:
        self._pool = pool
        self._ensure_schema()

    def _ensure_schema(self) -> None:
        try:
            with self._pool.connection() as conn:
                conn.execute(_CREATE_TABLE_POSTGRES)
                conn.commit()
        except Exception as exc:
            logger.warning(
                "QueryAnalyticsExportPostgresBackend: schema setup failed: %s", exc
            )

    def create_export(self, record: Dict[str, Any]) -> None:
        with self._pool.connection() as conn:
            conn.execute(_INSERT_SQL_POSTGRES, _build_record_tuple(record))
            conn.commit()

    def update_export(self, export_id: str, **fields: Any) -> None:
        if not fields:
            return
        _validate_update_fields(fields)
        set_clauses = [f"{k} = %s" for k in fields]
        params: List[Any] = list(fields.values()) + [export_id]
        with self._pool.connection() as conn:
            conn.execute(
                f"UPDATE query_analytics_exports SET {', '.join(set_clauses)} WHERE id = %s",
                params,
            )
            conn.commit()

    def list_exports(self, export_id: Optional[str] = None) -> List[Dict[str, Any]]:
        with self._pool.connection() as conn:
            if export_id is not None:
                rows = conn.execute(
                    "SELECT * FROM query_analytics_exports WHERE id = %s "
                    "ORDER BY created_at DESC",
                    (export_id,),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM query_analytics_exports ORDER BY created_at DESC"
                ).fetchall()
            return [
                {
                    "id": row[0],
                    "initiated_by": row[1],
                    "created_at": row[2],
                    "status": row[3],
                    "filter_summary": row[4],
                    "file_path": row[5],
                    "file_size_bytes": row[6],
                    "row_count": row[7],
                    "error_message": row[8],
                    "retention_until": row[9],
                }
                for row in rows
            ]

    def evict_old_exports(self, now_ts: float) -> int:
        with self._pool.connection() as conn:
            rows = conn.execute(
                "SELECT id, file_path FROM query_analytics_exports "
                "WHERE retention_until IS NOT NULL AND retention_until < %s",
                (now_ts,),
            ).fetchall()
            count = _delete_expired_rows(rows, conn, "%s")
            conn.commit()
            return count

    def reconcile_orphaned_exports(
        self,
        threshold_seconds: float = _DEFAULT_ORPHAN_THRESHOLD_SECONDS,
        error: str = "interrupted by worker restart",
    ) -> int:
        """Mark stale pending/running exports as failed (Bug #1228).

        Cluster-safe: only affects exports whose created_at is older than
        threshold_seconds ago, preserving legitimately-running exports on
        other cluster nodes.

        Returns the number of rows transitioned to 'failed'.
        """
        cutoff_ts = time.time() - threshold_seconds
        with self._pool.connection() as conn:
            cur = conn.execute(
                "UPDATE query_analytics_exports "
                "SET status = 'failed', error_message = %s "
                "WHERE status IN ('pending', 'running') AND created_at < %s",
                (error, cutoff_ts),
            )
            conn.commit()
            return int(cur.rowcount)


# ---------------------------------------------------------------------------
# Export service
# ---------------------------------------------------------------------------


class QueryAnalyticsExportService:
    """Service for exporting search event log data to Excel files.

    Manages export lifecycle: create pending record, fetch+write Excel,
    update completed/failed, and evict old files.
    """

    def __init__(self, backend: Any, golden_repos_dir: str) -> None:
        self._backend = backend
        self._golden_repos_dir = golden_repos_dir

    def export_path(self, export_id: str) -> Path:
        """Compute the export file path for a given export_id.

        Raises ValueError if export_id is not a valid UUID (prevents path traversal).
        """
        _validate_export_id(export_id)
        return (
            Path(self._golden_repos_dir)
            / "cidx-exports"
            / "query-analytics"
            / f"{export_id}.xlsx"
        )

    def ensure_export_dir(self) -> None:
        """Create the export directory if it does not exist."""
        export_dir = Path(self._golden_repos_dir) / "cidx-exports" / "query-analytics"
        export_dir.mkdir(parents=True, exist_ok=True)

    def build_filter_summary(self, filters: Dict[str, Any]) -> str:
        """Build a human-readable summary string from the applied filters."""
        parts = []

        if filters.get("user"):
            parts.append(f"user={filters['user']}")
        if filters.get("repo_alias"):
            parts.append(f"repo_alias={filters['repo_alias']}")
        if filters.get("search_type") and filters["search_type"] != "all":
            parts.append(f"search_type={filters['search_type']}")
        if filters.get("cache_hit_filter") and filters["cache_hit_filter"] != "all":
            parts.append(f"cache_hit={filters['cache_hit_filter']}")

        from_ts = filters.get("from_timestamp")
        to_ts = filters.get("to_timestamp")
        if from_ts is not None:
            dt = datetime.fromtimestamp(from_ts, tz=timezone.utc)
            parts.append(f"from={dt.strftime('%Y-%m-%d')}")
        if to_ts is not None:
            dt = datetime.fromtimestamp(to_ts, tz=timezone.utc)
            parts.append(f"to={dt.strftime('%Y-%m-%d')}")

        if not parts:
            return "All searches"
        return ", ".join(parts)

    def _fetch_rows(
        self, search_backend: Any, filters: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Fetch rows from search_event_log applying the given filters.

        Filters supported by the backend are passed directly. cache_hit_filter
        is applied post-query in Python (the backend has no such predicate).
        """
        username = filters.get("user") or None
        search_type = filters.get("search_type")
        if search_type == "all":
            search_type = None
        repo_alias = filters.get("repo_alias") or None
        from_ts = filters.get("from_timestamp")
        to_ts = filters.get("to_timestamp")

        events, _ = search_backend.query(
            username=username,
            search_type=search_type,
            repo_alias=repo_alias,
            from_ts=from_ts,
            to_ts=to_ts,
            limit=_EXPORT_LIMIT,
            offset=0,
        )

        cache_hit_filter = filters.get("cache_hit_filter", "all")
        if cache_hit_filter == "hits_only":
            events = [
                e
                for e in events
                if e.get("voyage_cache_hit") or e.get("cohere_cache_hit")
            ]
        elif cache_hit_filter == "misses_only":
            events = [
                e
                for e in events
                if not e.get("voyage_cache_hit") and not e.get("cohere_cache_hit")
            ]

        return cast(List[Dict[str, Any]], events)

    def run_export(
        self,
        export_id: str,
        filters: Dict[str, Any],
        initiated_by: str,
        export_retention_days: int,
        search_event_log_backend: Any,
    ) -> None:
        """Execute the export: fetch rows, write Excel, update DB record.

        Creates the DB record as 'pending', writes Excel, then updates to
        'completed' or 'failed'. Intended to be called by the BGM worker thread.
        """
        now = time.time()
        filter_summary = self.build_filter_summary(filters)
        record: Dict[str, Any] = {
            "id": export_id,
            "initiated_by": initiated_by,
            "created_at": now,
            "status": "pending",
            "filter_summary": filter_summary,
        }
        self._backend.create_export(record)

        try:
            rows = self._fetch_rows(search_event_log_backend, filters)
            self.ensure_export_dir()
            file_path = self.export_path(export_id)
            self._write_excel(rows, file_path)

            file_size = file_path.stat().st_size
            retention_until = now + export_retention_days * _SECONDS_PER_DAY
            self._backend.update_export(
                export_id,
                status="completed",
                file_path=str(file_path),
                file_size_bytes=file_size,
                row_count=len(rows),
                retention_until=retention_until,
            )
        except Exception as exc:
            logger.error(
                "QueryAnalyticsExportService: export %s failed: %s", export_id, exc
            )
            error_retention_until = now + export_retention_days * _SECONDS_PER_DAY
            self._backend.update_export(
                export_id,
                status="failed",
                error_message=str(exc),
                retention_until=error_retention_until,
            )

    def _write_excel(self, rows: List[Dict[str, Any]], file_path: Path) -> None:
        """Write rows to an Excel file at file_path with the canonical column order."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        # Write header row
        for col_idx, col_name in enumerate(_EXCEL_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data rows -- "user" in Excel maps to "username" in DB
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(_EXCEL_COLUMNS, start=1):
                db_key = "username" if col_name == "user" else col_name
                ws.cell(row=row_idx, column=col_idx, value=row.get(db_key))

        wb.save(str(file_path))

    def list_exports(self, export_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """Return export records, adding a computed 'download_link' field.

        For completed exports the link is the REST download endpoint.
        For all other statuses (pending, running, failed) the link is None.
        """
        records = self._backend.list_exports(export_id=export_id)
        for record in records:
            if record.get("status") == "completed":
                record["download_link"] = (
                    f"/api/admin/search-events/exports/{record['id']}/download"
                )
            else:
                record["download_link"] = None
        return cast(List[Dict[str, Any]], records)

    def evict_old_exports(self, now_ts: float) -> int:
        """Delegate eviction to the backend."""
        return cast(int, self._backend.evict_old_exports(now_ts=now_ts))

    def reconcile_orphaned_exports(
        self,
        threshold_seconds: float = _DEFAULT_ORPHAN_THRESHOLD_SECONDS,
        error: str = "interrupted by worker restart",
    ) -> int:
        """Reconcile orphaned exports: transition stale pending/running rows to failed.

        Delegates to the backend.  Cluster-safe: only affects exports whose
        created_at is older than threshold_seconds ago (default 300 s / 5 min).

        Called on server startup to clear exports left stuck by worker death,
        server restart, or infrastructure outage (Bug #1228).

        Returns the number of rows transitioned to 'failed'.
        """
        return cast(
            int,
            self._backend.reconcile_orphaned_exports(
                threshold_seconds=threshold_seconds,
                error=error,
            ),
        )
