"""Tests for QueryAnalyticsExportService (Story #1160)."""

import time
import uuid
from pathlib import Path

import pytest

from code_indexer.server.services.query_analytics_export_service import (
    QueryAnalyticsExportService,
    QueryAnalyticsExportSqliteBackend,
)
from code_indexer.server.services.search_event_log_writer import (
    SearchEventLogSqliteBackend,
)


@pytest.fixture
def tmp_db(tmp_path):
    db_path = str(tmp_path / "test.db")
    return db_path


@pytest.fixture
def export_backend(tmp_db):
    return QueryAnalyticsExportSqliteBackend(tmp_db)


@pytest.fixture
def search_backend(tmp_db):
    return SearchEventLogSqliteBackend(tmp_db)


@pytest.fixture
def export_dir(tmp_path):
    return tmp_path / "cidx-exports"


@pytest.fixture
def export_service(tmp_path, export_backend):
    return QueryAnalyticsExportService(
        backend=export_backend,
        golden_repos_dir=str(tmp_path),
    )


class TestQueryAnalyticsExportSqliteBackend:
    def test_create_and_list(self, export_backend):
        record = {
            "id": "test-id-1",
            "initiated_by": "alice",
            "created_at": time.time(),
            "status": "pending",
            "filter_summary": "All searches",
        }
        export_backend.create_export(record)
        exports = export_backend.list_exports()
        assert len(exports) == 1
        assert exports[0]["id"] == "test-id-1"
        assert exports[0]["status"] == "pending"
        assert exports[0]["initiated_by"] == "alice"

    def test_update_export(self, export_backend):
        record = {
            "id": "test-id-2",
            "initiated_by": "bob",
            "created_at": time.time(),
            "status": "pending",
            "filter_summary": "user=bob",
        }
        export_backend.create_export(record)
        export_backend.update_export("test-id-2", status="completed", row_count=42)
        exports = export_backend.list_exports(export_id="test-id-2")
        assert len(exports) == 1
        assert exports[0]["status"] == "completed"
        assert exports[0]["row_count"] == 42

    def test_list_by_id_empty(self, export_backend):
        exports = export_backend.list_exports(export_id="nonexistent")
        assert exports == []

    def test_evict_old_exports_removes_expired(self, export_backend, tmp_path):
        old_ts = time.time() - 1000
        old_file = tmp_path / "old.xlsx"
        old_file.write_bytes(b"fake")
        record = {
            "id": "old-id",
            "initiated_by": "carol",
            "created_at": old_ts - 1000,
            "status": "completed",
            "filter_summary": "All",
            "file_path": str(old_file),
            "retention_until": old_ts,
        }
        export_backend.create_export(record)
        count = export_backend.evict_old_exports(now_ts=time.time())
        assert count == 1
        assert not old_file.exists()
        exports = export_backend.list_exports()
        assert len(exports) == 0

    def test_evict_preserves_unexpired(self, export_backend):
        future_ts = time.time() + 86400
        record = {
            "id": "new-id",
            "initiated_by": "dave",
            "created_at": time.time(),
            "status": "completed",
            "filter_summary": "All",
            "retention_until": future_ts,
        }
        export_backend.create_export(record)
        count = export_backend.evict_old_exports(now_ts=time.time())
        assert count == 0
        exports = export_backend.list_exports()
        assert len(exports) == 1

    def test_evict_missing_file_no_error(self, export_backend):
        old_ts = time.time() - 1000
        record = {
            "id": "missing-file-id",
            "initiated_by": "eve",
            "created_at": old_ts - 1000,
            "status": "completed",
            "filter_summary": "All",
            "file_path": "/nonexistent/path/file.xlsx",
            "retention_until": old_ts,
        }
        export_backend.create_export(record)
        # Should not raise
        count = export_backend.evict_old_exports(now_ts=time.time())
        assert count == 1


class TestQueryAnalyticsExportServiceFiltering:
    def _insert_search_event(self, backend, **kwargs):
        """Helper to insert a search event with defaults."""
        from code_indexer.server.services.search_event_log_writer import (
            SearchEventRecord,
        )

        record = SearchEventRecord(
            timestamp=kwargs.get("timestamp", time.time()),
            username=kwargs.get("username", "testuser"),
            repo_alias=kwargs.get("repo_alias", "my-repo"),
            search_type=kwargs.get("search_type", "semantic"),
            query_text=kwargs.get("query_text", "test query"),
            voyage_cache_hit=kwargs.get("voyage_cache_hit", None),
            voyage_cache_mode=kwargs.get("voyage_cache_mode", None),
            voyage_latency_ms=kwargs.get("voyage_latency_ms", None),
            cohere_cache_hit=kwargs.get("cohere_cache_hit", None),
            cohere_cache_mode=kwargs.get("cohere_cache_mode", None),
            cohere_latency_ms=kwargs.get("cohere_latency_ms", None),
            total_latency_ms=kwargs.get("total_latency_ms", 100),
            result_count=kwargs.get("result_count", 5),
            node_id=kwargs.get("node_id", "node-1"),
            correlation_id=kwargs.get("correlation_id", None),
        )
        backend.insert_batch([record])

    def test_filter_by_user(self, export_service, search_backend):
        self._insert_search_event(search_backend, username="alice")
        self._insert_search_event(search_backend, username="bob")
        rows = export_service._fetch_rows(search_backend, {"user": "alice"})
        assert all(r["username"] == "alice" for r in rows)
        assert len(rows) == 1

    def test_filter_by_search_type(self, export_service, search_backend):
        self._insert_search_event(search_backend, search_type="semantic")
        self._insert_search_event(search_backend, search_type="fts")
        rows = export_service._fetch_rows(search_backend, {"search_type": "semantic"})
        assert len(rows) == 1
        assert rows[0]["search_type"] == "semantic"

    def test_filter_search_type_all_returns_all(self, export_service, search_backend):
        self._insert_search_event(search_backend, search_type="semantic")
        self._insert_search_event(search_backend, search_type="fts")
        rows = export_service._fetch_rows(search_backend, {"search_type": "all"})
        assert len(rows) == 2

    def test_filter_by_repo_alias(self, export_service, search_backend):
        self._insert_search_event(search_backend, repo_alias="repo-a")
        self._insert_search_event(search_backend, repo_alias="repo-b")
        rows = export_service._fetch_rows(search_backend, {"repo_alias": "repo-a"})
        assert len(rows) == 1

    def test_filter_by_timestamp_range(self, export_service, search_backend):
        now = time.time()
        self._insert_search_event(search_backend, timestamp=now - 200)
        self._insert_search_event(search_backend, timestamp=now - 100)
        self._insert_search_event(search_backend, timestamp=now)
        rows = export_service._fetch_rows(
            search_backend,
            {
                "from_timestamp": now - 150,
                "to_timestamp": now,
            },
        )
        assert len(rows) == 1

    def test_cache_hit_filter_hits_only(self, export_service, search_backend):
        self._insert_search_event(search_backend, voyage_cache_hit=True)
        self._insert_search_event(search_backend, voyage_cache_hit=False)
        self._insert_search_event(search_backend, cohere_cache_hit=True)
        rows = export_service._fetch_rows(
            search_backend, {"cache_hit_filter": "hits_only"}
        )
        assert len(rows) == 2
        for r in rows:
            assert r.get("voyage_cache_hit") or r.get("cohere_cache_hit")

    def test_cache_hit_filter_misses_only(self, export_service, search_backend):
        self._insert_search_event(search_backend, voyage_cache_hit=True)
        self._insert_search_event(
            search_backend, voyage_cache_hit=False, cohere_cache_hit=False
        )
        self._insert_search_event(search_backend)  # None/None
        rows = export_service._fetch_rows(
            search_backend, {"cache_hit_filter": "misses_only"}
        )
        assert len(rows) == 2

    def test_cache_hit_filter_all_returns_all(self, export_service, search_backend):
        self._insert_search_event(search_backend, voyage_cache_hit=True)
        self._insert_search_event(search_backend, voyage_cache_hit=False)
        rows = export_service._fetch_rows(search_backend, {"cache_hit_filter": "all"})
        assert len(rows) == 2


class TestQueryAnalyticsExportServiceExcel:
    def test_excel_correct_columns(self, export_service, search_backend, tmp_path):
        from code_indexer.server.services.search_event_log_writer import (
            SearchEventRecord,
        )

        record = SearchEventRecord(
            timestamp=time.time(),
            username="testuser",
            repo_alias="repo-x",
            search_type="semantic",
            query_text="hello world",
            voyage_cache_hit=True,
            voyage_cache_mode="on",
            voyage_latency_ms=50,
            cohere_cache_hit=None,
            cohere_cache_mode=None,
            cohere_latency_ms=None,
            total_latency_ms=120,
            result_count=10,
            node_id="node-1",
            correlation_id="corr-123",
        )
        search_backend.insert_batch([record])
        export_id = str(uuid.uuid4())
        export_service.run_export(
            export_id=export_id,
            filters={},
            initiated_by="admin",
            export_retention_days=30,
            search_event_log_backend=search_backend,
        )
        exports = export_service._backend.list_exports(export_id=export_id)
        assert len(exports) == 1
        assert exports[0]["status"] == "completed"
        assert exports[0]["row_count"] == 1
        file_path = exports[0]["file_path"]
        assert file_path is not None
        assert Path(file_path).exists()

        import openpyxl

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        expected = [
            "timestamp",
            "user",
            "repo_alias",
            "search_type",
            "query_text",
            "voyage_cache_hit",
            "voyage_cache_mode",
            "voyage_latency_ms",
            "cohere_cache_hit",
            "cohere_cache_mode",
            "cohere_latency_ms",
            "total_latency_ms",
            "result_count",
            "node_id",
            "correlation_id",
        ]
        assert headers == expected
        # Data row
        data_row = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
        assert data_row[1] == "testuser"  # user column

    def test_empty_export_header_only(self, export_service, search_backend):
        export_id = str(uuid.uuid4())
        export_service.run_export(
            export_id=export_id,
            filters={},
            initiated_by="admin",
            export_retention_days=30,
            search_event_log_backend=search_backend,
        )
        exports = export_service._backend.list_exports(export_id=export_id)
        assert exports[0]["status"] == "completed"
        assert exports[0]["row_count"] == 0
        import openpyxl

        wb = openpyxl.load_workbook(exports[0]["file_path"])
        ws = wb.active
        assert ws.max_row == 1  # header only

    def test_export_failure_sets_failed_status(self, export_service):
        """Pass a broken backend to trigger failure path."""

        class BrokenBackend:
            def query(self, **kwargs):
                raise RuntimeError("DB is broken")

        export_id = str(uuid.uuid4())
        export_service.run_export(
            export_id=export_id,
            filters={},
            initiated_by="admin",
            export_retention_days=30,
            search_event_log_backend=BrokenBackend(),
        )
        exports = export_service._backend.list_exports(export_id=export_id)
        assert exports[0]["status"] == "failed"
        assert exports[0]["error_message"] is not None


class TestServiceListExports:
    """Tests for QueryAnalyticsExportService.list_exports() with download_link."""

    def test_list_exports_empty(self, export_service):
        results = export_service.list_exports()
        assert results == []

    def test_completed_export_has_download_link(self, export_service, export_backend):
        record = {
            "id": "abc-completed",
            "initiated_by": "alice",
            "created_at": 1000.0,
            "status": "completed",
            "filter_summary": "All searches",
            "file_path": "/tmp/abc-completed.xlsx",
            "file_size_bytes": 512,
            "row_count": 10,
            "error_message": None,
            "retention_until": 2000.0,
        }
        export_backend.create_export(record)
        results = export_service.list_exports()
        assert len(results) == 1
        assert results[0]["download_link"] == (
            "/api/admin/search-events/exports/abc-completed/download"
        )

    def test_pending_export_has_no_download_link(self, export_service, export_backend):
        record = {
            "id": "abc-pending",
            "initiated_by": "alice",
            "created_at": 1000.0,
            "status": "pending",
            "filter_summary": "All searches",
        }
        export_backend.create_export(record)
        results = export_service.list_exports()
        assert results[0]["download_link"] is None

    def test_failed_export_has_no_download_link(self, export_service, export_backend):
        record = {
            "id": "abc-failed",
            "initiated_by": "alice",
            "created_at": 1000.0,
            "status": "failed",
            "filter_summary": "All searches",
            "error_message": "something broke",
        }
        export_backend.create_export(record)
        results = export_service.list_exports()
        assert results[0]["download_link"] is None

    def test_list_by_id(self, export_service, export_backend):
        for i, status in enumerate(["completed", "pending"]):
            export_backend.create_export(
                {
                    "id": f"id-{i}",
                    "initiated_by": "bob",
                    "created_at": float(i),
                    "status": status,
                    "filter_summary": "All",
                }
            )
        results = export_service.list_exports(export_id="id-0")
        assert len(results) == 1
        assert results[0]["id"] == "id-0"

    def test_list_by_unknown_id_empty(self, export_service):
        results = export_service.list_exports(export_id="nonexistent")
        assert results == []


class TestFilterSummary:
    def test_no_filters(self, export_service):
        summary = export_service.build_filter_summary({})
        assert summary == "All searches"

    def test_with_user(self, export_service):
        summary = export_service.build_filter_summary({"user": "alice"})
        assert "user=alice" in summary

    def test_with_search_type(self, export_service):
        summary = export_service.build_filter_summary({"search_type": "semantic"})
        assert "search_type=semantic" in summary

    def test_with_dates(self, export_service):
        # from_timestamp and to_timestamp should produce human-readable dates
        summary = export_service.build_filter_summary(
            {
                "from_timestamp": 1735689600.0,  # 2025-01-01
                "to_timestamp": 1748736000.0,  # 2025-06-01
            }
        )
        assert "from=" in summary
        assert "to=" in summary

    def test_multiple_filters(self, export_service):
        summary = export_service.build_filter_summary(
            {
                "user": "alice",
                "search_type": "semantic",
            }
        )
        assert "user=alice" in summary
        assert "search_type=semantic" in summary

    def test_export_path_construction(self, export_service, tmp_path):
        # export_path requires a valid UUID (prevents path traversal)
        valid_uuid = "12345678-1234-5678-1234-567812345678"
        path = export_service.export_path(valid_uuid)
        assert "cidx-exports" in str(path)
        assert "query-analytics" in str(path)
        assert f"{valid_uuid}.xlsx" in str(path)

    def test_export_path_rejects_non_uuid(self, export_service):
        with pytest.raises(ValueError, match="canonical UUID"):
            export_service.export_path("../../etc/passwd")
